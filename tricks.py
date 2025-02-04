import sqlite3
import string
import os
from openpyxl import load_workbook

def dict_to_sqlite(box_dict, table_name, database=str(os.path.abspath(os.path.dirname(__file__))+"/image_data.db")):
	conn = sqlite3.connect(database)
	c = conn.cursor()
	c.execute("drop table if exists " + table_name + ";")
	sql_statement = "create table if not exists " + table_name + " ("
	box_dict_list = list(box_dict.keys())
	table_columns = ""
	for i in range(0,len(box_dict_list),1):
		if(i<len(box_dict_list)-1):
			sql_statement += str(box_dict_list[i]) + ' text,'
			table_columns += str(box_dict_list[i]) + ', '
		else:
			sql_statement += str(box_dict_list[i]) + ' text)'
			table_columns += str(box_dict_list[i])
	c.execute(sql_statement)
	for i in range(0,len(box_dict[list(box_dict.keys())[1]])):
		instruction_row = []
		for j in box_dict_list:
			if(box_dict[j][i] is None):
				instruction_row.append("NULL")
			else:
				instruction_row.append(box_dict[j][i])
		values = ['(']
		[values.append('''"'''+str(i)+'''",''') for i in instruction_row]
		values.append(");") 
		values = "".join(values).replace(",)", ")")  # placeholder for values
		SQL = "INSERT INTO " + str(table_name) + "(" + str(table_columns) + ") values " + values
		#print(SQL)
		#print(instruction_row)
		c.execute(SQL)
		conn.commit()
	conn.close()

def update_in_table(table, update_column, value, conditional, print_indic=False):
	'''arbitrarily update stuff in the database
	'''
	if value is not None:
		value = "'" + value + "'"
	else:
		value = "null"
	SQL = """
		update """ + table + """
		set """ + update_column + """ = """ + value + """
		""" + """ """ + conditional + """;"""
	if(print_indic):
		print(SQL)
	run_SQL(SQL,'y')

def run_SQL(SQL, commit_indic='n', database=str(os.path.abspath(os.path.dirname(__file__))+"/image_data.db")):
	conn = sqlite3.connect(database)
	conn.row_factory = sqlite3.Row
	c = conn.cursor()
	c.execute(SQL)
	data = c.fetchall()
	if(commit_indic=='y'):
		conn.commit()
		return
	c.close()
	conn.close()
	data = [dict(i) for i in data]
	return data

def most_frequent(List):
	return max(set(List), key = List.count)

def send_to_excel(path, data, workbook, sheet="Sheet1", clear_indic='n', cells=None, headings='n'):
	excel_file = os.path.abspath(path + "/" + workbook + ".xlsx") #file to send data to
	#print(excel_file)
	wb = load_workbook(filename = excel_file)
	output_sheet = wb[sheet] #sheet to send data to

	#iterate through dict of lists, send to excel--indexes start at 1
	if cells==None:
		max_col = str(list(string.ascii_uppercase)[len(data)-1])
		clear_col = str(list(string.ascii_uppercase)[12])
		if clear_indic=='y':
			for row in output_sheet['A2:'+clear_col+'4269']:
				for cell in row:
					cell.value = None
		columns = list(data.keys())
		for i in range(0,len(columns),1):
			name = columns[i] #name in dict
			for j in range(0,len(data[columns[i]]),1):
				output_sheet.cell(row=j+2,column=i+1).value = data[name][j]
		if(headings=='y'):
			counter = 0
			for row in output_sheet['A1:'+clear_col+str(1)]:
				for cell in row:
					cell.value = None
			for row in output_sheet['A1:'+max_col+str(1)]:
				for cell in row:
					cell.value = columns[counter]
					counter+=1
	else:
		#logic for writing to a range of cells
		cells = cells.split(":")
		rows = [int("".join([i for i in i if i.isnumeric()])) for i in cells]
		cols = ["".join([i for i in i if not i.isnumeric()]) for i in cells]
		if clear_indic=='y':
			for row in output_sheet[cols[0]+str(rows[0])+":"+cols[1]+'4269']:
				for cell in row:
					cell.value = None
		cols = [int(ord(i.lower())-96) for i in cols]
		columns = list(data.keys())
		for i in range(0,cols[1]-cols[0]+1,1):
			for j in range(0,rows[1]-rows[0]+1,1):
				if(j==0):
					output_sheet.cell(row=j+rows[0],column=i+cols[0]).value = columns[i]
				else:
					output_sheet.cell(row=j+rows[0],column=i+cols[0]).value = data[columns[i]][j-1]
	wb.save(excel_file)

def list_to_dict(data_list):
	keys = list(data_list[0].keys())
	data_dict = {i: [] for i in keys}
	for i in range(0,len(data_list)):
		for j in keys:
			data_dict[j].append(data_list[i][j])
	return data_dict

def format_SQL(SQL):
    SQL = " ".join(SQL.splitlines())
    SQL = " ".join([i.strip().lower() for i in SQL.split() if len(i)>0])
    SQL = [str(i.strip().lower()) for i in SQL.split()]
    for i in range(0,len(SQL)):
        if SQL[i]=='union':
            SQL[i] = '\nunion\n'
        elif SQL[i]=='from':
            SQL[i] = '\nfrom'
        elif SQL[i]=='join':
            SQL[i+2] = SQL[i+2]+"\n"
        elif SQL[i]=='inner':
            SQL[i] = '\ninner'
        elif SQL[i]=='outer':
            SQL[i] = '\nouter'
        elif SQL[i]=='from':
            SQL[i] = '\nfrom'
        elif SQL[i]=='select':
            SQL[i] = '\nselect'
        elif SQL[i]=='(select':
            SQL[i] = '\n(select'
        elif SQL[i]=='where':
            SQL[i] = '\nwhere'
        elif SQL[i]=='group':
            SQL[i] =  "\ngroup"
        elif SQL[i]=='order':
            SQL[i] = "\norder"
        elif SQL[i]=='and':
            SQL[i] = '\nand'
        elif SQL[i]=='on':
            SQL[i] = '\non'
    for i in range(0,len(SQL)):
        if(i<len(SQL)-2):
            if (SQL[i].strip().lower()=='from' and SQL[i+1].strip().lower()=='(' and SQL[i+2].strip().lower() == 'select') or (SQL[i].strip().lower()=='from' and SQL[i+1].strip().lower()=='(select'):
                SQL[i] = '\nfrom'
                SQL[i+1] = '(select'
                SQL[i+2] = ''
    SQL = " ".join(SQL)
    SQL = os.linesep.join([s for s in SQL.splitlines() if s.strip()])
    SQL = SQL.replace(") )","))")
    return SQL